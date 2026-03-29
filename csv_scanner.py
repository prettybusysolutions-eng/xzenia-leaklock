#!/usr/bin/env python3
"""
CSV Scanner — LeakLock
Accepts CSV or Excel (.xlsx) billing/invoice files, auto-detects columns,
applies all 15 leak patterns. Returns structured results per pattern.
"""

import io
import csv
import re
import uuid
import json
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s'
)
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ─── Chunked processing constant ─────────────────────────────────────────────
CHUNK_SIZE = 5000

# ─── Column auto-detection mappings ──────────────────────────────────────────

COLUMN_PATTERNS = {
    'amount': [
        r'amount', r'revenue', r'charge', r'price', r'total', r'invoice.?amount',
        r'payment.?amount', r'value', r'fee', r'cost', r'sale'
    ],
    'date': [
        r'date', r'invoice.?date', r'created', r'issued', r'transaction.?date',
        r'billed', r'period', r'timestamp', r'payment.?date'
    ],
    'status': [
        r'status', r'payment.?status', r'state', r'result', r'outcome'
    ],
    'customer_id': [
        r'customer', r'client', r'account', r'customer.?id', r'client.?id',
        r'account.?id', r'contact', r'company'
    ],
    'description': [
        r'description', r'item', r'product', r'service', r'memo', r'note',
        r'details', r'line.?item', r'category'
    ],
    'discount': [
        r'discount', r'coupon', r'promo', r'reduction', r'credit.?memo',
        r'adjustment', r'deduction'
    ],
}


def normalize_header(h: str) -> str:
    """Lowercase, strip, collapse spaces."""
    return re.sub(r'\s+', ' ', h.strip().lower())


def detect_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    Map canonical field names → column index.
    Returns dict like {'amount': 2, 'date': 0, 'status': 3, ...}
    None means not found.
    """
    mapping = {k: None for k in COLUMN_PATTERNS}
    norm_headers = [normalize_header(h) for h in headers]

    for canonical, patterns in COLUMN_PATTERNS.items():
        for i, nh in enumerate(norm_headers):
            for pat in patterns:
                if re.search(pat, nh):
                    mapping[canonical] = i
                    break
            if mapping[canonical] is not None:
                break

    return mapping


def parse_amount(val: str) -> Optional[float]:
    """
    Parse messy amount strings: $1,234.56 → 1234.56
    Returns None on failure.
    """
    if not val or not val.strip():
        return None
    # Remove currency symbols, whitespace
    cleaned = re.sub(r'[$€£¥\s]', '', str(val).strip())
    # Remove commas used as thousands separators
    cleaned = cleaned.replace(',', '')
    # Handle parentheses as negatives: (100) → -100
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return None


# Alias used by new patterns
_parse_amount = parse_amount


def parse_date(val: str) -> Optional[datetime]:
    """Try multiple date formats, return datetime or None."""
    if not val or not val.strip():
        return None
    formats = [
        '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d',
        '%m-%d-%Y', '%d-%m-%Y',
        '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S',
        '%b %d, %Y', '%B %d, %Y', '%d %b %Y',
        '%m/%d/%y', '%d/%m/%y'
    ]
    val = val.strip()
    for fmt in formats:
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


# Alias used by new patterns
_parse_date = parse_date


def parse_status(val: str) -> str:
    """Normalize status to paid/failed/pending/unknown."""
    if not val:
        return 'unknown'
    v = val.strip().lower()
    if any(x in v for x in ['paid', 'success', 'complete', 'settled', 'captured', 'cleared']):
        return 'paid'
    if any(x in v for x in ['fail', 'declined', 'error', 'bounced', 'reject', 'chargeback']):
        return 'failed'
    if any(x in v for x in ['pend', 'open', 'draft', 'unpaid', 'due', 'outstanding']):
        return 'pending'
    if any(x in v for x in ['refund', 'void', 'credit', 'reversed']):
        return 'refund'
    return 'unknown'


def parse_csv(raw_bytes: bytes) -> Tuple[List[Dict], Dict[str, Optional[int]], int]:
    """
    Parse raw CSV bytes into normalized row dicts.
    Returns (rows, column_mapping, total_raw_rows)
    """
    # Try UTF-8, fallback to latin-1
    try:
        text = raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw_bytes.decode('latin-1')

    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    if not all_rows:
        return [], {}, 0

    # Find the header row (first non-empty row)
    header_idx = 0
    for i, row in enumerate(all_rows):
        if any(cell.strip() for cell in row):
            header_idx = i
            break

    headers = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1:]
    total_raw = len(data_rows)

    col_map = detect_columns(headers)

    rows = []
    for raw_row in data_rows:
        if not any(cell.strip() for cell in raw_row):
            continue  # skip blank lines

        def get(col):
            idx = col_map.get(col)
            if idx is None or idx >= len(raw_row):
                return None
            return raw_row[idx].strip() or None

        amount_raw = get('amount')
        amount = parse_amount(amount_raw) if amount_raw else None
        
        date_raw = get('date')
        date = parse_date(date_raw) if date_raw else None
        
        status_raw = get('status')
        status = parse_status(status_raw) if status_raw else 'unknown'
        
        discount_raw = get('discount')
        discount = parse_amount(discount_raw) if discount_raw else None

        rows.append({
            'amount': amount,
            'date': date,
            'status': status,
            'customer_id': get('customer_id'),
            'description': get('description'),
            'discount': discount,
            '_raw': raw_row,
        })

    return rows, col_map, total_raw


# ─── Pattern detectors (CSV-adapted) ─────────────────────────────────────────

def detect_discount_drift(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 1: Discount Drift — discounts increasing over time."""
    discount_rows = [r for r in rows if r.get('discount') and r.get('discount', 0) > 0 and r.get('date')]
    if len(discount_rows) < 3:
        # Also check if many rows have no amount but have a description with discount keywords
        discount_keyword_rows = [r for r in rows if r.get('description') and 
                                  any(k in (r.get('description') or '').lower() for k in ['discount', 'promo', 'coupon', 'off'])]
        if len(discount_keyword_rows) < 2:
            return None
        total_discount_est = sum((r.get('amount') or 0) * 0.15 for r in discount_keyword_rows)
        affected = len(discount_keyword_rows)
        return {
            'pattern': 'discount_drift',
            'pattern_name': 'Discount Drift',
            'description': 'Discounts/promotions detected in billing data — may indicate discount creep',
            'amount_estimate': int(total_discount_est),
            'confidence': 0.55,
            'severity': 'medium',
            'affected_rows': affected,
            'details': {'discount_keyword_rows': affected}
        }

    # Sort by date
    discount_rows.sort(key=lambda r: r['date'])
    # Get discount amounts over time periods
    mid = len(discount_rows) // 2
    early = [r['discount'] for r in discount_rows[:mid] if r['discount']]
    late = [r['discount'] for r in discount_rows[mid:] if r['discount']]
    
    if not early or not late:
        return None
    
    early_avg = sum(early) / len(early)
    late_avg = sum(late) / len(late)
    
    if late_avg > early_avg * 1.1:  # Discounts grew >10%
        total_discount = sum(r['discount'] for r in discount_rows)
        drift_pct = (late_avg - early_avg) / early_avg
        return {
            'pattern': 'discount_drift',
            'pattern_name': 'Discount Drift',
            'description': f'Average discount grew {drift_pct:.0%} over time — leaking margin',
            'amount_estimate': int(total_discount * drift_pct),
            'confidence': 0.72,
            'severity': 'medium',
            'affected_rows': len(discount_rows),
            'details': {
                'early_avg_discount': round(early_avg, 2),
                'late_avg_discount': round(late_avg, 2),
                'drift_pct': round(drift_pct * 100, 1),
                'rows_with_discounts': len(discount_rows)
            }
        }
    return None


def detect_usage_underbilling(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 2: Usage Underbilling — usage items without charges."""
    usage_rows = [r for r in rows if r.get('description') and 
                  any(k in (r.get('description') or '').lower() for k in 
                      ['usage', 'metered', 'overage', 'unit', 'seat', 'per-user', 'add-on'])]
    
    if not usage_rows:
        return None
    
    # Look for usage rows with zero or no amount
    underbilled = [r for r in usage_rows if not r.get('amount') or r.get('amount', 0) == 0]
    
    if not underbilled:
        # Check if usage amounts seem low relative to other rows
        all_amounts = [r.get('amount', 0) for r in rows if r.get('amount')]
        usage_amounts = [r.get('amount', 0) for r in usage_rows if r.get('amount')]
        
        if all_amounts and usage_amounts:
            overall_avg = sum(all_amounts) / len(all_amounts)
            usage_avg = sum(usage_amounts) / len(usage_amounts)
            if usage_avg < overall_avg * 0.5:
                gap = (overall_avg - usage_avg) * len(usage_rows)
                return {
                    'pattern': 'usage_underbilling',
                    'pattern_name': 'Usage Underbilling',
                    'description': 'Usage-based charges appear below expected rates',
                    'amount_estimate': int(gap * 0.5),
                    'confidence': 0.55,
                    'severity': 'high',
                    'affected_rows': len(usage_rows),
                    'details': {'usage_rows': len(usage_rows), 'avg_gap': round(overall_avg - usage_avg, 2)}
                }
        return None
    
    # Estimate underbilling
    paid_usage = [r for r in usage_rows if r.get('amount')]
    avg_usage_charge = sum(r['amount'] for r in paid_usage) / len(paid_usage) if paid_usage else 100
    estimated_loss = avg_usage_charge * len(underbilled)
    
    return {
        'pattern': 'usage_underbilling',
        'pattern_name': 'Usage Underbilling',
        'description': f'{len(underbilled)} usage/metered items with no charge captured',
        'amount_estimate': int(estimated_loss),
        'confidence': 0.78,
        'severity': 'high',
        'affected_rows': len(underbilled),
        'details': {
            'unbilled_usage_rows': len(underbilled),
            'avg_usage_charge': round(avg_usage_charge, 2)
        }
    }


def detect_renewal_erosion(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 3: Renewal Erosion — cancels/downgrades at renewal."""
    cancel_keywords = ['cancel', 'churn', 'terminate', 'downgrade', 'pause', 'suspend']
    cancel_rows = [r for r in rows if r.get('description') and 
                   any(k in (r.get('description') or '').lower() for k in cancel_keywords)]
    
    if not cancel_rows:
        return None
    
    lost_revenue = sum(r.get('amount', 0) or 0 for r in cancel_rows)
    
    return {
        'pattern': 'renewal_erosion',
        'pattern_name': 'Renewal Erosion',
        'description': f'{len(cancel_rows)} cancellations/downgrades found in billing data',
        'amount_estimate': int(lost_revenue),
        'confidence': 0.80,
        'severity': 'high',
        'affected_rows': len(cancel_rows),
        'details': {
            'cancellation_rows': len(cancel_rows),
            'total_lost': lost_revenue
        }
    }


def detect_scope_creep(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 4: Scope Creep — services delivered, no invoice."""
    service_keywords = ['feature', 'custom', 'consulting', 'implementation', 'setup', 
                        'onboard', 'training', 'support', 'service', 'project']
    service_rows = [r for r in rows if r.get('description') and 
                    any(k in (r.get('description') or '').lower() for k in service_keywords)]
    
    if not service_rows:
        return None
    
    # Service rows with no amount = potential scope creep
    unbilled_services = [r for r in service_rows if not r.get('amount') or r.get('amount', 0) == 0]
    
    if not unbilled_services:
        return None
    
    # Estimate based on avg of billed service rows
    billed_services = [r for r in service_rows if r.get('amount', 0) and r['amount'] > 0]
    avg_service = (sum(r['amount'] for r in billed_services) / len(billed_services)) if billed_services else 500
    
    estimated_loss = avg_service * len(unbilled_services)
    
    return {
        'pattern': 'scope_creep',
        'pattern_name': 'Scope Creep',
        'description': f'{len(unbilled_services)} service line items with no amount billed',
        'amount_estimate': int(estimated_loss),
        'confidence': 0.60,
        'severity': 'medium',
        'affected_rows': len(unbilled_services),
        'details': {
            'unbilled_service_rows': len(unbilled_services),
            'billed_service_rows': len(billed_services),
            'avg_service_charge': round(avg_service, 2)
        }
    }


def detect_credit_memo_overissuance(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 5: Credit Memo Overissuance — excessive credits."""
    credit_keywords = ['credit', 'refund', 'adjustment', 'credit memo', 'reversal']
    credit_rows = [r for r in rows if 
                   (r.get('description') and any(k in (r.get('description') or '').lower() for k in credit_keywords)) or
                   (r.get('status') == 'refund')]
    
    if len(credit_rows) < 2:
        return None
    
    total_credits = sum(abs(r.get('amount') or 0) for r in credit_rows)
    paid_rows = [r for r in rows if r.get('status') == 'paid' and r.get('amount')]
    total_revenue = sum(r['amount'] for r in paid_rows)
    
    if total_revenue == 0:
        # Can't determine ratio without revenue data
        return None
    
    credit_ratio = total_credits / total_revenue
    
    if credit_ratio > 0.08:  # >8% credit ratio is suspicious
        excess = total_credits - (total_revenue * 0.03)  # 3% credit is normal
        return {
            'pattern': 'credit_memo_overissuance',
            'pattern_name': 'Credit Memo Overissuance',
            'description': f'Credit ratio is {credit_ratio:.1%} — above healthy 3% threshold',
            'amount_estimate': int(max(0, excess)),
            'confidence': 0.68,
            'severity': 'medium',
            'affected_rows': len(credit_rows),
            'details': {
                'credit_rows': len(credit_rows),
                'total_credits': round(total_credits, 2),
                'credit_ratio_pct': round(credit_ratio * 100, 1),
                'total_revenue': round(total_revenue, 2)
            }
        }
    return None


def detect_failed_payment_neglect(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 6: Failed Payment Neglect — failed charges never retried."""
    failed_rows = [r for r in rows if r.get('status') == 'failed' and r.get('amount')]
    
    if not failed_rows:
        return None
    
    total_failed = sum(r['amount'] or 0 for r in failed_rows)
    
    # Check if any failed payments were subsequently recovered
    # (simplified: look for a paid row with same amount or customer)
    recovered_count = 0
    for fr in failed_rows:
        if fr.get('customer_id'):
            matching = [r for r in rows if 
                        r.get('customer_id') == fr['customer_id'] and 
                        r.get('status') == 'paid' and 
                        r.get('amount') == fr.get('amount')]
            if matching:
                recovered_count += 1
    
    unrecovered = len(failed_rows) - recovered_count
    unrecovered_total = total_failed * (unrecovered / len(failed_rows)) if failed_rows else 0
    
    return {
        'pattern': 'failed_payment_neglect',
        'pattern_name': 'Failed Payment Neglect',
        'description': f'{unrecovered} failed payments with no recovery detected (${unrecovered_total:,.0f})',
        'amount_estimate': int(unrecovered_total),
        'confidence': 0.90,
        'severity': 'high',
        'affected_rows': unrecovered,
        'details': {
            'total_failed': len(failed_rows),
            'recovered': recovered_count,
            'unrecovered': unrecovered,
            'total_failed_amount': round(total_failed, 2)
        }
    }


def detect_contract_term_amnesia(rows: List[Dict]) -> Optional[Dict]:
    """Pattern 7: Contract Term Amnesia — stale pricing, no rate updates."""
    dated_rows = [r for r in rows if r.get('date') and r.get('amount') and r.get('customer_id')]
    
    if len(dated_rows) < 6:
        return None
    
    # Group by customer
    by_customer = {}
    for r in dated_rows:
        cid = r['customer_id']
        by_customer.setdefault(cid, []).append(r)
    
    stale_customers = []
    for cid, crows in by_customer.items():
        if len(crows) < 3:
            continue
        crows.sort(key=lambda r: r['date'])
        # Check if pricing is completely flat over 6+ months
        dates = [r['date'] for r in crows]
        date_span = (max(dates) - min(dates)).days
        if date_span < 180:
            continue
        
        amounts = [r['amount'] for r in crows if r.get('amount')]
        if not amounts:
            continue
        min_a, max_a = min(amounts), max(amounts)
        # If amounts vary by <2% across 6+ months, likely no price updates
        if max_a > 0 and (max_a - min_a) / max_a < 0.02:
            stale_customers.append({
                'customer_id': cid,
                'avg': sum(amounts) / len(amounts),
                'span_days': date_span,
                'rows': len(crows)
            })
    
    if not stale_customers:
        return None
    
    # Estimate: 5% annual rate increase they're missing
    estimated_loss = sum(c['avg'] * 0.05 * (c['span_days'] / 365) * c['rows'] 
                         for c in stale_customers)
    
    return {
        'pattern': 'contract_term_amnesia',
        'pattern_name': 'Contract Term Amnesia',
        'description': f'{len(stale_customers)} customers with no price changes over 6+ months — may be on stale rates',
        'amount_estimate': int(estimated_loss),
        'confidence': 0.50,
        'severity': 'low',
        'affected_rows': sum(c['rows'] for c in stale_customers),
        'details': {
            'stale_customers': len(stale_customers),
            'avg_span_days': round(sum(c['span_days'] for c in stale_customers) / len(stale_customers))
        }
    }


# ─── BAR / KAVA LOUNGE PATTERNS ───────────────────────────────────────────────

def _find_column(rows, candidates):
    """Find column key from a list of candidate names (returns key string or None)."""
    if not rows:
        return None
    sample = rows[0]
    for c in candidates:
        if c in sample:
            return c
    return None


def detect_ghost_inventory(rows: List[Dict]) -> Optional[Dict]:
    """
    Detects potential shrinkage where pours aren't logged.
    """
    try:
        amount_idx = _find_column(rows, ['amount', 'total', 'subtotal', 'price', 'line_total'])
        qty_idx = _find_column(rows, ['quantity', 'qty', 'count', 'units'])
        
        if amount_idx is None or qty_idx is None:
            return None
        
        ghost_pours = []
        for r in rows:
            amt = _parse_amount(r.get(amount_idx, 0))
            qty = _parse_amount(r.get(qty_idx, 1))
            if amt == 0 and qty and qty > 0:
                ghost_pours.append(r)
        
        if not ghost_pours:
            return None
        
        estimated_loss = len(ghost_pours) * 8  # Assume $8 avg drink price
        
        return {
            'pattern': 'ghost_inventory',
            'pattern_name': 'Ghost Inventory',
            'description': f'{len(ghost_pours)} items logged with zero price — possible untracked pours or comps',
            'amount_estimate': int(estimated_loss),
            'confidence': 0.55,
            'severity': 'medium',
            'affected_rows': len(ghost_pours),
            'details': {'suspicious_items': len(ghost_pours)}
        }
    except Exception:
        return None


def detect_tab_leakage(rows: List[Dict]) -> Optional[Dict]:
    """
    Detects open tabs that were never closed/paid.
    """
    try:
        status_idx = _find_column(rows, ['status', 'state', 'payment_status', 'txn_status'])
        
        if status_idx is None:
            return None
        
        open_tabs = []
        for r in rows:
            status = str(r.get(status_idx, '')).lower() if status_idx else ''
            if 'open' in status or 'pending' in status or 'unpaid' in status:
                open_tabs.append(r)
        
        if not open_tabs:
            return None
        
        estimated_loss = len(open_tabs) * 25  # avg tab $25
        
        return {
            'pattern': 'tab_leakage',
            'pattern_name': 'Tab Leakage',
            'description': f'{len(open_tabs)} open/unpaid tabs — customers who walked without paying',
            'amount_estimate': int(estimated_loss),
            'confidence': 0.60,
            'severity': 'high',
            'affected_rows': len(open_tabs),
            'details': {'open_tabs': len(open_tabs)}
        }
    except Exception:
        return None


def detect_event_deposit_loss(rows: List[Dict]) -> Optional[Dict]:
    """
    Detects event deposits that were never captured or refunded.
    """
    try:
        event_idx = _find_column(rows, ['event', 'event_name', 'booking', 'description'])
        status_idx = _find_column(rows, ['status', 'state', 'booking_status', 'event_status'])
        amount_idx = _find_column(rows, ['amount', 'deposit', 'fee', 'charge', 'total'])
        
        if event_idx is None:
            return None
        
        no_shows = []
        for r in rows:
            status = str(r.get(status_idx, '')).lower() if status_idx else ''
            if 'no-show' in status or 'forfeit' in status or 'no show' in status:
                no_shows.append(r)
        
        if not no_shows:
            return None
        
        estimated_loss = sum(
            _parse_amount(r.get(amount_idx, 50)) or 50 for r in no_shows
        )
        
        return {
            'pattern': 'event_deposit_loss',
            'pattern_name': 'Event Deposit Loss',
            'description': f'{len(no_shows)} no-show events — deposits forfeited but not captured',
            'amount_estimate': int(estimated_loss),
            'confidence': 0.65,
            'severity': 'high',
            'affected_rows': len(no_shows),
            'details': {'no_show_events': len(no_shows)}
        }
    except Exception:
        return None


def detect_membership_freeze(rows: List[Dict]) -> Optional[Dict]:
    """
    Detects members on auto-billing who stopped coming but are still being charged.
    """
    try:
        customer_idx = _find_column(rows, ['customer', 'member', 'name', 'email'])
        amount_idx = _find_column(rows, ['amount', 'charge', 'fee', 'subscription_amount'])
        date_idx = _find_column(rows, ['date', 'last_activity', 'last_visit', 'transaction_date'])
        
        if customer_idx is None:
            return None
        
        customer_activity = {}
        for r in rows:
            cust = str(r.get(customer_idx, '')).lower()
            if not cust:
                continue
            dt = _parse_date(str(r.get(date_idx, ''))) if date_idx else None
            amt = _parse_amount(str(r.get(amount_idx, 0))) or 0 if amount_idx else 0
            
            if cust not in customer_activity:
                customer_activity[cust] = {'last_date': dt, 'total_charged': 0}
            
            if dt and (not customer_activity[cust]['last_date'] or dt > customer_activity[cust]['last_date']):
                customer_activity[cust]['last_date'] = dt
            customer_activity[cust]['total_charged'] += amt
        
        frozen_members = []
        for cust, data in customer_activity.items():
            if data['total_charged'] > 0 and data['last_date']:
                days_since = (datetime.now() - data['last_date']).days
                if days_since > 60 and data['total_charged'] > 0:
                    frozen_members.append({'customer': cust, 'days_inactive': days_since, 'charged': data['total_charged']})
        
        if not frozen_members:
            return None
        
        estimated_loss = sum(m['charged'] * 0.5 for m in frozen_members)
        
        return {
            'pattern': 'membership_freeze',
            'pattern_name': 'Membership Freeze',
            'description': f'{len(frozen_members)} members on auto-bill with no activity in 60+ days',
            'amount_estimate': int(estimated_loss),
            'confidence': 0.55,
            'severity': 'medium',
            'affected_rows': len(frozen_members),
            'details': {'frozen_members': len(frozen_members)}
        }
    except Exception:
        return None


# ─── NEW PATTERNS (12-15) ─────────────────────────────────────────────────────

def detect_duplicate_billing(rows, col_map=None):
    """Pattern 12: Detect duplicate charges: same customer, same amount, within 7 days."""
    # Support both dict-row (new) and col_map-based (legacy) calling conventions
    if col_map is not None:
        amt_col = col_map.get('amount')
        date_col = col_map.get('date')
        cust_col = col_map.get('customer')
        if not amt_col or not date_col or not cust_col:
            return None
        def get_amt(r): return _parse_amount(str(r.get(amt_col, '') or ''))
        def get_date(r): return _parse_date(str(r.get(date_col, '') or ''))
        def get_cust(r): return str(r.get(cust_col, '') or '').strip()
    else:
        # Use pre-parsed dict rows from parse_csv
        def get_amt(r): return r.get('amount')
        def get_date(r): return r.get('date')
        def get_cust(r): return str(r.get('customer_id') or '').strip()

    seen = {}
    duplicates = []
    total = 0.0

    for row in rows:
        try:
            amt = get_amt(row)
            date = get_date(row)
            cust = get_cust(row)
            if not amt or not date or not cust:
                continue

            key = (cust, round(amt, 2))
            if key in seen:
                prev_date = seen[key]
                delta = abs((date - prev_date).days)
                if delta <= 7:
                    duplicates.append(row)
                    total += amt
            else:
                seen[key] = date
        except Exception:
            continue

    if len(duplicates) >= 2:
        return {
            'pattern': 'duplicate_billing',
            'pattern_name': 'Duplicate Billing',
            'description': f'{len(duplicates)} potential duplicate charges detected within 7-day windows (${total:,.0f})',
            'amount_estimate': int(total),
            'confidence': 0.80,
            'severity': 'high',
            'affected_rows': len(duplicates),
            'details': {'duplicate_count': len(duplicates), 'total_amount': total}
        }
    return None


def detect_refund_abuse(rows, col_map=None):
    """Pattern 13: Detect refund rate exceeding 15% of revenue."""
    if col_map is not None:
        amt_col = col_map.get('amount')
        status_col = col_map.get('status')
        desc_col = col_map.get('description')
        if not amt_col:
            return None
        def get_amt(r): return _parse_amount(str(r.get(amt_col, '') or ''))
        def get_status(r): return str(r.get(status_col, '') or '').lower() if status_col else ''
        def get_desc(r): return str(r.get(desc_col, '') or '').lower() if desc_col else ''
    else:
        def get_amt(r): return r.get('amount')
        def get_status(r): return str(r.get('status') or '').lower()
        def get_desc(r): return str(r.get('description') or '').lower()

    total_charged = 0.0
    total_refunded = 0.0
    refund_count = 0

    refund_keywords = {'refund', 'refunded', 'return', 'returned', 'reversal', 'chargeback', 'credit back'}

    for row in rows:
        try:
            amt = get_amt(row)
            if not amt:
                continue

            status = get_status(row)
            desc = get_desc(row)

            is_refund = any(kw in status or kw in desc for kw in refund_keywords)

            if is_refund or amt < 0:
                total_refunded += abs(amt)
                refund_count += 1
            else:
                total_charged += abs(amt)
        except Exception:
            continue

    if total_charged > 0 and total_refunded > 0:
        refund_rate = total_refunded / total_charged
        if refund_rate > 0.15:
            return {
                'pattern': 'refund_abuse',
                'pattern_name': 'Refund Abuse',
                'description': f'Refund rate is {refund_rate:.0%} of revenue — industry norm is <5%. {refund_count} refunds totaling ${total_refunded:,.0f}',
                'amount_estimate': int(total_refunded * 0.6),  # recoverable portion
                'confidence': 0.72,
                'severity': 'high',
                'affected_rows': refund_count,
                'details': {'refund_rate': refund_rate, 'total_refunded': total_refunded, 'total_charged': total_charged}
            }
    return None


def detect_pricing_inconsistency(rows, col_map=None):
    """Pattern 14: Detect same service billed at different prices to different customers."""
    if col_map is not None:
        amt_col = col_map.get('amount')
        desc_col = col_map.get('description')
        if not amt_col or not desc_col:
            return None
        def get_amt(r): return _parse_amount(str(r.get(amt_col, '') or ''))
        def get_desc(r): return str(r.get(desc_col, '') or '').strip().lower()
    else:
        def get_amt(r): return r.get('amount')
        def get_desc(r): return str(r.get('description') or '').strip().lower()

    service_prices = {}

    for row in rows:
        try:
            amt = get_amt(row)
            desc = get_desc(row)
            if not amt or not desc or amt <= 0:
                continue

            key = desc[:30]
            if key not in service_prices:
                service_prices[key] = []
            service_prices[key].append(amt)
        except Exception:
            continue

    inconsistencies = []
    total_gap = 0.0

    for service, prices in service_prices.items():
        if len(prices) < 3:
            continue
        min_p = min(prices)
        max_p = max(prices)
        if min_p > 0 and (max_p - min_p) / min_p > 0.30:  # >30% variance
            gap = (max_p - min_p) * len(prices)
            inconsistencies.append((service, min_p, max_p, len(prices)))
            total_gap += gap

    if inconsistencies:
        return {
            'pattern': 'pricing_inconsistency',
            'pattern_name': 'Pricing Inconsistency',
            'description': f'{len(inconsistencies)} services billed at inconsistent rates — price variance >30% for same service',
            'amount_estimate': int(total_gap * 0.2),  # conservative recoverable
            'confidence': 0.65,
            'severity': 'medium',
            'affected_rows': sum(i[3] for i in inconsistencies),
            'details': {'inconsistent_services': len(inconsistencies), 'total_variance': total_gap}
        }
    return None


def detect_late_invoice(rows, col_map=None):
    """Pattern 15: Detect invoices paid >60 days late — cash flow leakage."""
    if col_map is not None:
        date_col = col_map.get('date')
        amt_col = col_map.get('amount')
        status_col = col_map.get('status')
        if not date_col or not amt_col:
            return None
        def get_date(r): return _parse_date(str(r.get(date_col, '') or ''))
        def get_amt(r): return _parse_amount(str(r.get(amt_col, '') or ''))
    else:
        def get_date(r): return r.get('date')
        def get_amt(r): return r.get('amount')

    dates = []

    for row in rows:
        try:
            d = get_date(row)
            amt = get_amt(row)
            if not d or not amt:
                continue
            dates.append(d)
        except Exception:
            continue

    if len(dates) < 5:
        return None

    dates_sorted = sorted(dates)
    gaps = []
    for i in range(1, len(dates_sorted)):
        gap = (dates_sorted[i] - dates_sorted[i - 1]).days
        if gap > 60:
            gaps.append(gap)

    if len(gaps) >= 2:
        avg_gap = sum(gaps) / len(gaps)
        all_amts = [get_amt(r) or 0 for r in rows]
        avg_monthly = sum(all_amts) / max(1, len(dates)) * 30
        estimated_loss = avg_monthly * (avg_gap / 30) * len(gaps) * 0.05

        if estimated_loss > 100:
            return {
                'pattern': 'late_invoice',
                'pattern_name': 'Late Invoice Gaps',
                'description': f'{len(gaps)} billing gaps >60 days detected. Average gap: {avg_gap:.0f} days. Cash flow cost estimated at ${estimated_loss:,.0f}',
                'amount_estimate': int(estimated_loss),
                'confidence': 0.60,
                'severity': 'medium',
                'affected_rows': len(gaps),
                'details': {'gap_count': len(gaps), 'avg_gap_days': avg_gap}
            }
    return None


# ─── Chunked detector runner ──────────────────────────────────────────────────

def _run_detectors(rows: List[Dict]) -> List[Dict]:
    """Run all detectors against a list of pre-parsed rows."""
    detectors = [
        detect_discount_drift,
        detect_usage_underbilling,
        detect_renewal_erosion,
        detect_scope_creep,
        detect_credit_memo_overissuance,
        detect_failed_payment_neglect,
        detect_contract_term_amnesia,
        detect_ghost_inventory,
        detect_tab_leakage,
        detect_event_deposit_loss,
        detect_membership_freeze,
        detect_duplicate_billing,
        detect_refund_abuse,
        detect_pricing_inconsistency,
        detect_late_invoice,
    ]
    results = []
    for det in detectors:
        try:
            r = det(rows)
            if r:
                results.append(r)
        except Exception as e:
            logger.warning(f'Detector {det.__name__} error: {e}')
    return results


def _run_detectors_chunked(df_rows: List[Dict]) -> List[Dict]:
    """Run detectors in chunks for large datasets."""
    if len(df_rows) <= CHUNK_SIZE:
        return _run_detectors(df_rows)

    chunks = [df_rows[i:i + CHUNK_SIZE] for i in range(0, len(df_rows), CHUNK_SIZE)]
    all_leaks = {}

    for chunk in chunks:
        chunk_leaks = _run_detectors(chunk)
        for leak in chunk_leaks:
            pname = leak.get('pattern_name', '')
            if pname not in all_leaks:
                all_leaks[pname] = leak
            else:
                # Accumulate amounts across chunks
                all_leaks[pname]['amount_estimate'] = (
                    all_leaks[pname].get('amount_estimate', 0) +
                    leak.get('amount_estimate', 0)
                )
                # Keep highest confidence
                if leak.get('confidence', 0) > all_leaks[pname].get('confidence', 0):
                    all_leaks[pname]['confidence'] = leak['confidence']

    return list(all_leaks.values())


# ─── Excel support ────────────────────────────────────────────────────────────

def _scan_csv_bytes(raw_bytes: bytes) -> dict:
    """
    Core CSV scan: parse bytes → run 15 detectors → return results dict.
    Used internally by scan_csv() and _scan_excel().
    """
    scan_id = str(uuid.uuid4())
    errors = []

    try:
        rows, col_map, total_raw = parse_csv(raw_bytes)
    except Exception as e:
        return {
            'scan_id': scan_id,
            'rows_parsed': 0,
            'total_revenue': 0,
            'total_leakage': 0,
            'patterns_triggered': 0,
            'leaks': [],
            'column_mapping': {},
            'errors': [f'CSV parse error: {str(e)}']
        }

    if not rows:
        return {
            'scan_id': scan_id,
            'rows_parsed': 0,
            'total_revenue': 0,
            'total_leakage': 0,
            'patterns_triggered': 0,
            'leaks': [],
            'column_mapping': col_map,
            'errors': ['No data rows found in CSV']
        }

    paid_rows = [r for r in rows if r.get('status') in ('paid', 'unknown') and r.get('amount') and r['amount'] > 0]
    total_revenue = sum(r['amount'] for r in paid_rows)

    # Use chunked runner for large datasets
    leaks = _run_detectors_chunked(rows)

    for leak in leaks:
        leak['scan_id'] = scan_id

    total_leakage = sum(l['amount_estimate'] for l in leaks)
    leaks.sort(key=lambda l: l['amount_estimate'], reverse=True)

    return {
        'scan_id': scan_id,
        'rows_parsed': len(rows),
        'total_raw_rows': total_raw,
        'total_revenue': round(total_revenue, 2),
        'total_leakage': round(total_leakage, 2),
        'patterns_triggered': len(leaks),
        'leaks': leaks,
        'column_mapping': {k: v for k, v in col_map.items() if v is not None},
        'errors': errors,
        'scanned_at': datetime.now().isoformat()
    }


def _scan_excel(raw_bytes: bytes) -> dict:
    """Scan Excel (.xlsx) file — lazy import openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return {
            'scan_id': str(uuid.uuid4()),
            'rows_parsed': 0,
            'total_revenue': 0,
            'total_leakage': 0,
            'patterns_triggered': 0,
            'leaks': [],
            'error': 'openpyxl not installed. Run: pip install openpyxl',
            'column_mapping': {}
        }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {
            'scan_id': str(uuid.uuid4()),
            'rows_parsed': 0,
            'total_revenue': 0,
            'total_leakage': 0,
            'patterns_triggered': 0,
            'leaks': [],
            'error': f'Could not open Excel file: {str(e)}',
            'column_mapping': {}
        }

    all_results = []
    total_rows = 0
    total_revenue = 0.0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(c).strip() if c is not None else '' for c in rows[0]]
        data_rows = rows[1:]

        if not data_rows:
            continue

        # Convert to CSV bytes for existing scanner
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(['' if c is None else str(c) for c in row])

        csv_bytes = buf.getvalue().encode('utf-8')
        sheet_result = _scan_csv_bytes(csv_bytes)

        total_rows += sheet_result.get('rows_parsed', 0)
        total_revenue += sheet_result.get('total_revenue', 0)
        all_results.append(sheet_result)

    wb.close()

    if not all_results:
        return {
            'scan_id': str(uuid.uuid4()),
            'rows_parsed': 0,
            'total_revenue': 0,
            'total_leakage': 0,
            'patterns_triggered': 0,
            'leaks': [],
            'column_mapping': {}
        }

    # Deduplicate leaks by pattern_name, keep highest amount
    seen_patterns = {}
    for r in all_results:
        for leak in r.get('leaks', []):
            pname = leak.get('pattern_name', '')
            if pname not in seen_patterns or leak.get('amount_estimate', 0) > seen_patterns[pname].get('amount_estimate', 0):
                seen_patterns[pname] = leak
    all_leaks = list(seen_patterns.values())

    total_leakage = sum(l.get('amount_estimate', 0) for l in all_leaks)

    return {
        'scan_id': str(uuid.uuid4()),
        'rows_parsed': total_rows,
        'total_revenue': total_revenue,
        'total_leakage': total_leakage,
        'patterns_triggered': len(all_leaks),
        'leaks': sorted(all_leaks, key=lambda x: x.get('amount_estimate', 0), reverse=True),
        'column_mapping': all_results[0].get('column_mapping', {}) if all_results else {}
    }


def _scan_excel_legacy(raw_bytes: bytes) -> dict:
    """Old .xls format — not supported, return helpful message."""
    return {
        'scan_id': str(uuid.uuid4()),
        'rows_parsed': 0,
        'total_revenue': 0,
        'total_leakage': 0,
        'patterns_triggered': 0,
        'leaks': [],
        'error': 'Old .xls format detected. Please save as .xlsx or .csv and re-upload.',
        'column_mapping': {}
    }


# ─── Main scan function ───────────────────────────────────────────────────────

def scan_csv(raw_bytes: bytes) -> Dict:
    """
    Auto-detect format (CSV or Excel) and scan for revenue leaks.
    Runs all 15 detectors. Returns structured results.
    """
    # Detect file format from magic bytes
    if raw_bytes[:4] in (b'PK\x03\x04', b'PK\x05\x06', b'PK\x07\x08'):
        # ZIP-based format = Excel (.xlsx)
        return _scan_excel(raw_bytes)
    elif raw_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        # OLE2 compound = Legacy Excel (.xls)
        return _scan_excel_legacy(raw_bytes)
    else:
        # Default: CSV (or TSV, plain text)
        return _scan_csv_bytes(raw_bytes)


def save_scan_to_db(scan_result: Dict, db_config: Dict) -> bool:
    """Save scan results to saas_detected_leaks table."""
    try:
        import psycopg2
        import json as _json
        
        conn = psycopg2.connect(**db_config)
        try:
            with conn.cursor() as cur:
                scan_id = scan_result['scan_id']
                for leak in scan_result['leaks']:
                    cur.execute("""
                        INSERT INTO saas_detected_leaks
                        (scan_id, pattern, amount_estimate, confidence, severity, details, status)
                        VALUES (%s, %s, %s, %s, %s, %s, 'detected')
                    """, (
                        scan_id,
                        leak.get('pattern', leak.get('pattern_name', 'unknown')),
                        leak['amount_estimate'],
                        leak['confidence'],
                        leak['severity'],
                        _json.dumps({
                            **leak.get('details', {}),
                            'pattern_name': leak.get('pattern_name', ''),
                            'description': leak.get('description', ''),
                            'affected_rows': leak.get('affected_rows', 0),
                        })
                    ))
            conn.commit()
            return True
        finally:
            conn.close()
    except Exception as e:
        logger.error(f'DB save error: {e}')
        return False


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('Usage: python csv_scanner.py <csv_or_xlsx_file>')
        sys.exit(1)
    
    with open(sys.argv[1], 'rb') as f:
        raw = f.read()
    
    result = scan_csv(raw)
    print(json.dumps(result, indent=2, default=str))
